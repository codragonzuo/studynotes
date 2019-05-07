# Usage: python perf2excel.py perfpara.xml
import sys,time
from openpyxl import Workbook
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
print(sys.argv)
#tree = ET.parse("perfpara.xml")
tree = ET.parse(sys.argv[1])
root = tree.getroot()
count = 1
excel=Workbook()
sheet=excel.active
sheet.cell(row=1,column=1).value='NAME'
sheet.cell(row=1,column=2).value='MIN'
sheet.cell(row=1,column=3).value='MAX'
sheet.cell(row=1,column=4).value='RANGE'
sheet.cell(row=1,column=5).value=sys.argv[1]
for record in root.findall('record'):
    count=count+1
#    print(record[0].text.strip(),record[1].text.strip(),record[3].text.strip(),record[2].text.strip())
    sheet.cell(row=count,column=1).value=record[0].text.strip()
    sheet.cell(row=count,column=2).value=record[1].text.strip()
    sheet.cell(row=count,column=3).value=record[3].text.strip()
    sheet.cell(row=count,column=4).value=record[2].text.strip()
print(count)
filename = 'result-'+ time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time())) + '.xlsx'
excel.save(filename)
